#!/usr/bin/env python3
"""
Sistema de Extracción de Facturas con Gemini AI
- API Key editable en tiempo real desde la UI
- Fallback automático: gemini-2.5-flash → gemini-2.0-flash
- Analiza TODAS las páginas del PDF
- Tabla de cuentas → RUBRO incorporada
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import fitz  # PyMuPDF
import json
import re
import io
import os
import time
import traceback
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

try:
    from google import genai
    from google.genai import types
    GEMINI_OK = True
except ImportError:
    GEMINI_OK = False

app = Flask(__name__, static_folder='static')
CORS(app)

# ─────────────────────────────────────────────────────────
# CONFIGURACIÓN GLOBAL — modificable en tiempo real
# ─────────────────────────────────────────────────────────
CONFIG = {
    "api_key": os.environ.get("GEMINI_API_KEY", "AIzaSyBUodKbKBABuEOd69QI9gASnOwoPlHYF58"),
    # Modelos en orden de preferencia (fallback automático)
    "models": ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-2.0-flash-lite"],
    "max_retries": 3,      # Reintentos por modelo
    "retry_delay": 20,     # Segundos entre reintentos por rate-limit
}

# ─────────────────────────────────────────────────────────
# TABLA DE CUENTAS → RUBRO
# ─────────────────────────────────────────────────────────
TABLA_CUENTAS = {
    "52-1-08-01-01": "BAÑOS",
    "52-1-08-01-02": "RIEGO",
    "52-1-08-01-03": "FUENTES",
    "52-1-08-01-04": "EQUIPOS DE SENTINA",
    "52-1-08-01-05": "EQUIPOS CONTRA INCENDIO",
    "52-1-08-01-06": "INSTALACIONES ELECTRICAS",
    "52-1-08-01-07": "TRANSFORMADORES Y SUBESTACIONES",
    "52-1-08-01-08": "EQUIPOS DE RADIO",
    "52-1-08-01-09": "VEHICULOS SKYJACK",
    "52-1-08-01-10": "CISTERNAS",
    "52-1-08-01-11": "SISTEMA DE AGUA POTABLE",
    "52-1-08-01-12": "AIRE ACONDICIONADO",
    "52-1-08-01-13": "ESCALERAS Y ELEVADORES",
    "52-1-08-01-14": "HERRAMIENTAS",
    "52-1-08-01-15": "AUDIO",
    "52-1-08-01-16": "CENTRAL TELEFONICA",
    "52-1-08-01-17": "CAMARAS DE SEGURIDAD",
    "52-1-08-01-18": "GENERACION",
    "52-1-08-01-19": "UPS CENTRAL",
    "52-1-08-01-20": "PUERTAS ELECTRICAS",
    "52-1-08-01-21": "MEDIDORES DE AGUA",
    "52-1-08-01-22": "VALLAS",
    "52-1-08-01-23": "VIDEO PATIO DE COMIDA",
    "52-1-08-01-24": "LOCALES VACIOS Y NUEVAS INSTALACIONES",
    "52-1-08-01-25": "SISTEMA DE GAS",
    "52-1-08-01-26": "VOZ Y DATOS",
    "52-1-08-01-27": "IMPREVISTOS",
    "52-1-24-02-01": "FOCOS Y LAMPARAS",
    "52-1-08-02-04": "AGUA POTABLE",
    "41-2-01-02-06": "ANTENAS",
    "41-2-01-02-07": "MANTENIMIENTO AC",
    "43-5-01-01-05": "REEMBOLSO DE GASTOS",
    "52-1-18-01-07": "TASA RECOLECCION DE BASURA",
    "52-1-28-01-01": "SUMINISTROS DE OFICINA",
    "52-1-18-01-04": "INTERAGUA",
    "10-2-01-03-01": "CAPEX (O ACTIVOS EN PROCESO)",
    "10-2-02-05-01": "CONSTRUCCIONES EN PROCESO",
    "52-2-04-01-05": "VARIOS",
    "52-4-02-01-07": "GASTOS INTERMEDIARIOS",
    "52-1-18-01-03": "PLANILLAS",
    "52-1-28-03-13": "OTROS GASTOS ADMINISTRATIVOS",
    "10-1-03-04-01": "DIESEL/INVENTARIO INSUMOS",
    "52-1-28-07-09": "PROTECTORES MRN",
    "52-2-11-03-03": "IMPRESIONES PUBLICIDAD CLIENTES",
    "52-2-12-01-09": "GASTOS PANORAMIX",
    "52-1-28-02-05": "LIMPIEZA AMPLIACION",
    "52-1-08-01-29": "MANTENIMIENTOS TECNICOS AMPLIACION",
    "52-1-08-02-32": "MANTENIMIENTOS OPERATIVOS AMPLIACION",
    "52-1-24-02-11": "INSUMOS AMPLIACION",
    "52-1-18-01-09": "ENERGIA AMPLIACION",
    "52-1-18-01-10": "AGUA POTABLE AMPLIACION",
    "52-2-12-01-01": "ACTIVACIONES EXPERIENCIA",
    "52-1-28-02-01": "TECNICO",
}

TABLA_TEXTO = "\n".join(f"  {k} → {v}" for k, v in TABLA_CUENTAS.items())


# ─────────────────────────────────────────────────────────
# PROMPT PRINCIPAL
# ─────────────────────────────────────────────────────────
def build_prompt():
    return f"""Eres un experto en extracción de datos de facturas comerciales ecuatorianas.
Se te entregan TODAS LAS PÁGINAS del documento en orden (factura en pág.1, orden de compra en pág.2-3, etc.).

TAREA: Extrae los campos y devuelve ÚNICAMENTE un objeto JSON válido (sin markdown, sin ```, sin texto extra).

{{
  "fecha": "Fecha de EMISIÓN de la factura. Formato DD/MM/AAAA. Busca 'Fecha Emisión' o 'Fecha y hora de Autorización'. Solo la fecha de la factura, no de la orden.",
  "proveedor": "Nombre COMERCIAL de quien EMITE la factura (tiene su RUC y logo arriba a la izquierda). NUNCA pongas 'MOBILSOL', 'INMOBILIARIA DEL SOL' ni variantes — esas somos NOSOTROS el comprador.",
  "factura": "Número completo de factura. Formato: 001-001-000002702.",
  "solped": "SOLPED o REQUISICIÓN. BUSCA EN TODAS LAS PÁGINAS. Patrón: SOL seguido de 7+ dígitos (ej: SOL0012163). Lugares: campo 'REQUISICIÓN' en la OC, campo 'OBSERVACIONES', 'Información Adicional'. Si NO encuentras → ''",
  "orden": "Número de Orden de Compra. BUSCA EN TODAS LAS PÁGINAS. Patrón: OC seguido de 7+ dígitos (ej: OC0015541). Si no existe → ''",
  "rubro": "Elige el rubro MÁS APROPIADO según qué se compra. Devuelve SOLO el nombre exacto de esta tabla:\n{TABLA_TEXTO}",
  "cuenta": "Código contable que corresponde al rubro elegido (ej: 52-1-08-01-12). Si no estás seguro → ''",
  "proyecto": "SOLO si el documento dice explícitamente 'PROYECTO:' o 'Proyecto:' seguido de nombre. Si no aparece → ''",
  "descripcion": "Descripción del trabajo/material. Busca primero en OBSERVACIONES de la OC, luego en ítems de la factura. Elige la más completa.",
  "sub_total": "Monto subtotal sin IVA. Solo número decimal. Ej: 27.53",
  "iva": "Monto del IVA. Solo número decimal. Ej: 4.13. Si es 0 → 0.00",
  "total": "Total de la factura. Solo número decimal. Ej: 31.66"
}}

REGLAS CRÍTICAS:
1. PROVEEDOR = quien factura (parte superior de la factura con su RUC). MOBILSOL = comprador → NUNCA proveedor.
2. SOLPED: busca en TODAS las páginas. En OC está en campo "REQUISICIÓN".
3. ORDEN OC: busca en TODAS las páginas. Puede estar en la factura o en la OC.
4. Ejemplos de RUBRO: ventilador/AC → AIRE ACONDICIONADO (52-1-08-01-12); focos → FOCOS Y LAMPARAS (52-1-24-02-01); generador → GENERACION (52-1-08-01-18).
5. PROYECTO: vacío si no dice literalmente "Proyecto:" en el documento.
6. Datos no encontrados → vacío "". NUNCA inventes valores.
7. Responde SOLO el JSON, sin ningún texto adicional."""


# ─────────────────────────────────────────────────────────
# FUNCIONES DE PROCESAMIENTO
# ─────────────────────────────────────────────────────────

def pdf_to_images(pdf_bytes, zoom=2.0):
    """Convierte todas las páginas del PDF a JPEG."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    mat = fitz.Matrix(zoom, zoom)
    for i, page in enumerate(doc):
        pix = page.get_pixmap(matrix=mat)
        jpg = pix.tobytes("jpeg")
        images.append(jpg)
        print(f"    Pág {i+1}: {len(jpg)//1024} KB")
    n = len(doc)
    doc.close()
    return images, n


def clean_json(raw):
    """Extrae JSON puro de la respuesta."""
    text = raw.strip()
    text = re.sub(r'^```(?:json)?\s*', '', text, flags=re.MULTILINE)
    text = re.sub(r'\s*```\s*$', '', text, flags=re.MULTILINE)
    start = text.find('{')
    end   = text.rfind('}')
    if start != -1 and end != -1:
        text = text[start:end+1]
    return text.strip()


def normalize(data):
    """Normaliza y valida los datos extraídos."""
    required = ["fecha", "proveedor", "factura", "solped", "orden",
                "rubro", "cuenta", "proyecto", "descripcion", "sub_total", "iva", "total"]
    for f in required:
        if f not in data or data[f] is None:
            data[f] = ""
        else:
            data[f] = str(data[f]).strip()

    # Bloquear MOBILSOL como proveedor
    prov = data.get("proveedor", "").lower()
    if any(x in prov for x in ["mobilsol", "movilsol", "inmobiliaria del sol"]):
        data["proveedor"] = ""

    # Normalizar SOLPED
    s = data.get("solped", "")
    if s and not re.match(r'^SOL\d+$', s, re.IGNORECASE):
        m = re.search(r'SOL\d{5,10}', s, re.IGNORECASE)
        data["solped"] = m.group(0).upper() if m else ""

    # Normalizar ORDEN
    o = data.get("orden", "")
    if o and not re.match(r'^OC\d+$', o, re.IGNORECASE):
        m = re.search(r'OC\d{5,10}', o, re.IGNORECASE)
        data["orden"] = m.group(0).upper() if m else ""

    # Sincronizar RUBRO ↔ CUENTA
    rubro = data.get("rubro", "").upper().strip()
    cuenta = data.get("cuenta", "").strip()
    if cuenta and cuenta in TABLA_CUENTAS:
        data["rubro"] = TABLA_CUENTAS[cuenta]
    elif rubro:
        for c, d in TABLA_CUENTAS.items():
            if d.upper() == rubro:
                data["cuenta"] = c
                data["rubro"] = d
                break

    # Normalizar montos
    for f in ["sub_total", "iva", "total"]:
        val = data.get(f, "").replace("$", "").replace(",", "").strip()
        try:
            float(val)
        except (ValueError, TypeError):
            val = ""
        data[f] = val

    return data


def extract_with_gemini(pdf_bytes, api_key=None):
    """
    Extrae datos usando Gemini con fallback automático de modelos.
    Intenta: gemini-2.5-flash → gemini-2.0-flash → gemini-2.0-flash-lite
    """
    if not GEMINI_OK:
        return None, "Librería google-genai no instalada. Ejecute: pip install google-genai"

    key = api_key or CONFIG["api_key"]
    if not key:
        return None, "API Key no configurada. Ingrese su API Key en la casilla superior."

    client = genai.Client(api_key=key)
    prompt = build_prompt()

    print(f"  Convirtiendo PDF a imágenes...")
    images, num_pages = pdf_to_images(pdf_bytes)
    if not images:
        return None, "No se pudieron extraer imágenes del PDF"

    print(f"  {num_pages} página(s) extraídas")

    # Construir partes (imágenes + prompt)
    base_parts = [types.Part.from_bytes(data=img, mime_type="image/jpeg") for img in images]
    base_parts.append(types.Part.from_text(text=prompt))

    last_error = "Error desconocido"
    models_tried = []

    # Intentar cada modelo disponible
    for model in CONFIG["models"]:
        models_tried.append(model)
        max_retries = CONFIG["max_retries"]

        for attempt in range(1, max_retries + 1):
            try:
                print(f"  [{model}] Intento {attempt}/{max_retries}...")
                t0 = time.time()

                response = client.models.generate_content(
                    model=model,
                    contents=base_parts,
                    config=types.GenerateContentConfig(temperature=0.1)
                )

                elapsed = time.time() - t0
                raw = response.text.strip()
                print(f"  [{model}] OK en {elapsed:.1f}s | {len(raw)} chars")
                print(f"  Preview: {raw[:200]}")

                # Parsear JSON
                clean = clean_json(raw)
                data = json.loads(clean)
                data = normalize(data)
                data["_paginas"] = num_pages
                data["_model"] = model
                return data, None

            except json.JSONDecodeError as e:
                last_error = f"Error parseando JSON de Gemini: {e}"
                print(f"  [{model}] JSON error: {e}")
                break  # No reintentar si el JSON es inválido

            except Exception as e:
                err_str = str(e)
                last_error = err_str

                if "RESOURCE_EXHAUSTED" in err_str or "429" in err_str:
                    if attempt < max_retries:
                        wait = CONFIG["retry_delay"] * attempt
                        print(f"  [{model}] Rate-limit, esperando {wait}s...")
                        time.sleep(wait)
                    else:
                        print(f"  [{model}] Cuota agotada, probando siguiente modelo...")
                        break  # Pasar al siguiente modelo

                elif "PERMISSION_DENIED" in err_str or "403" in err_str:
                    return None, "API Key inválida o sin permisos para este modelo."

                elif "NOT_FOUND" in err_str or "404" in err_str:
                    print(f"  [{model}] Modelo no disponible, probando siguiente...")
                    break  # Pasar al siguiente modelo

                else:
                    print(f"  [{model}] Error: {err_str[:200]}")
                    if attempt >= max_retries:
                        break

    return None, (
        f"Todos los modelos agotaron su cuota o fallaron. "
        f"Modelos probados: {', '.join(models_tried)}. "
        f"Último error: {last_error[:200]}. "
        f"Espere unos minutos y reintente, o use una nueva API Key."
    )


# ─────────────────────────────────────────────────────────
# GENERACIÓN DE EXCEL
# ─────────────────────────────────────────────────────────

def create_excel(registros):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control Facturas"

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    norm_fill = PatternFill("solid", fgColor="FFFFFF")
    warn_fill = PatternFill("solid", fgColor="FFFBEB")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    headers = ["FECHA","PROVEEDOR","FACTURA","SOLPED","ORDEN",
               "RUBRO","CUENTA","PROYECTO","DESCRIPCIÓN DEL TRABAJO",
               "SUB TOTAL","IVA","TOTAL FACTURAR"]
    widths  = [12, 30, 22, 14, 14, 30, 20, 28, 55, 12, 10, 14]
    campos  = ["fecha","proveedor","factura","solped","orden",
               "rubro","cuenta","proyecto","descripcion",
               "sub_total","iva","total"]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill      = hdr_fill
        c.font      = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 35

    for ri, reg in enumerate(registros, 2):
        fill = alt_fill if ri % 2 == 0 else norm_fill
        for ci, campo in enumerate(campos, 1):
            val = reg.get(campo, "")
            c   = ws.cell(row=ri, column=ci, value=val)
            if not val and campo in ("fecha", "proveedor", "factura", "solped"):
                c.fill = warn_fill
            else:
                c.fill = fill
            c.border    = thin
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 9))
        ws.row_dimensions[ri].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────
# RUTAS API
# ─────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_file("static/index.html")


@app.route("/api/health")
def health():
    key = CONFIG["api_key"]
    return jsonify({
        "status": "ok",
        "models": CONFIG["models"],
        "model": CONFIG["models"][0],
        "gemini_ready": GEMINI_OK,
        "api_key_set": bool(key),
        "api_key": key,
        "api_key_preview": (key[:8] + "..." + key[-4:]) if key else "",
        "cuentas": len(TABLA_CUENTAS),
    })


@app.route("/api/config", methods=["GET"])
def get_config():
    key = CONFIG["api_key"]
    return jsonify({
        "model": CONFIG["models"][0],
        "models": CONFIG["models"],
        "api_key_set": bool(key),
        "api_key": key,
        "api_key_preview": (key[:8] + "..." + key[-4:]) if key else "",
        "cuentas": len(TABLA_CUENTAS),
    })


@app.route("/api/config", methods=["POST"])
def set_config():
    """Actualiza la API Key en tiempo real."""
    try:
        body = request.get_json()
        if not body:
            return jsonify({"error": "Body vacío"}), 400

        new_key = body.get("api_key", "").strip()
        if not new_key:
            return jsonify({"error": "API Key vacía"}), 400
        if not new_key.startswith("AIza"):
            return jsonify({"error": "Formato inválido (debe iniciar con 'AIza')"}), 400

        CONFIG["api_key"] = new_key
        print(f"[CONFIG] API Key actualizada → {new_key[:8]}...{new_key[-4:]}")

        return jsonify({
            "success": True,
            "message": "API Key actualizada correctamente",
            "api_key_preview": new_key[:8] + "..." + new_key[-4:],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/test-key", methods=["POST"])
def test_key():
    """Prueba si una API Key es válida."""
    try:
        body = request.get_json() or {}
        key = body.get("api_key", "").strip() or CONFIG["api_key"]

        if not key:
            return jsonify({"valid": False, "error": "Sin API Key"}), 400
        if not GEMINI_OK:
            return jsonify({"valid": False, "error": "google-genai no instalado"}), 500

        client = genai.Client(api_key=key)
        # Usar el primer modelo disponible para la prueba
        model = CONFIG["models"][0]
        resp = client.models.generate_content(
            model=model,
            contents="Responde solo: OK",
            config=types.GenerateContentConfig(temperature=0, max_output_tokens=5)
        )
        return jsonify({
            "valid": True,
            "message": f"API Key válida ✓  Modelo {model} disponible",
            "response": resp.text.strip()[:20],
        })
    except Exception as e:
        err = str(e)
        if "PERMISSION_DENIED" in err or "403" in err:
            msg = "API Key inválida o sin permisos"
        elif "RESOURCE_EXHAUSTED" in err or "429" in err:
            msg = "Cuota agotada (key válida, pero sin cuota libre hoy)"
        elif "NOT_FOUND" in err or "404" in err:
            msg = "Modelo no encontrado"
        else:
            msg = err[:160]
        return jsonify({"valid": False, "error": msg}), 400


@app.route("/api/set-key", methods=["POST"])
def set_key_alias():
    """
    Guarda y verifica la API Key. Compatible con el frontend existente.
    Responde con {success, message, warning, api_key_preview}
    """
    try:
        body = request.get_json() or {}
        new_key = body.get("api_key", "").strip()
        if not new_key:
            return jsonify({"error": "API Key vacía"}), 400
        if not new_key.startswith("AIza"):
            return jsonify({"error": "Formato inválido (debe iniciar con 'AIza')"}), 400

        CONFIG["api_key"] = new_key
        print(f"[SET-KEY] → {new_key[:8]}...{new_key[-4:]}")

        # Intentar verificar la key
        warning = None
        if GEMINI_OK:
            try:
                client = genai.Client(api_key=new_key)
                client.models.generate_content(
                    model=CONFIG["models"][0],
                    contents="OK",
                    config=types.GenerateContentConfig(temperature=0, max_output_tokens=3)
                )
            except Exception as e:
                err = str(e)
                if "RESOURCE_EXHAUSTED" in err or "429" in err:
                    warning = "Key válida pero cuota agotada hoy. Se usará mañana o con nueva key."
                elif "PERMISSION_DENIED" in err or "403" in err:
                    return jsonify({"error": "API Key inválida o sin permisos"}), 400

        return jsonify({
            "success": True,
            "message": "API Key guardada" + (" y verificada ✓" if not warning else ""),
            "warning": warning,
            "api_key_preview": new_key[:8] + "..." + new_key[-4:],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cuentas")
def get_cuentas():
    return jsonify(TABLA_CUENTAS)


@app.route("/api/extract", methods=["POST"])
def extract_factura():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No se recibió ningún archivo"}), 400

        file = request.files["file"]
        if not file.filename:
            return jsonify({"error": "Nombre de archivo vacío"}), 400
        if not file.filename.lower().endswith(".pdf"):
            return jsonify({"error": "Solo se aceptan archivos PDF"}), 400

        # API Key del request tiene prioridad (soporta form data Y header)
        api_key_override = (
            request.form.get("api_key", "").strip() or
            request.headers.get("X-Gemini-Key", "").strip() or
            None
        )
        # Si la UI envió una key y es distinta a la actual, actualizarla globalmente
        if api_key_override and api_key_override.startswith("AIza"):
            CONFIG["api_key"] = api_key_override

        pdf_bytes = file.read()
        if not pdf_bytes:
            return jsonify({"error": "Archivo PDF vacío"}), 400

        print(f"\n{'='*60}")
        print(f"[EXTRACT] {file.filename} ({len(pdf_bytes)//1024} KB)")
        key_used = api_key_override or CONFIG["api_key"]
        print(f"  API Key: {key_used[:8]}...{key_used[-4:]}")

        data, error = extract_with_gemini(pdf_bytes, api_key=api_key_override)
        if error:
            return jsonify({"error": error}), 500

        print(f"[OK] proveedor={data.get('proveedor')} | factura={data.get('factura')} | solped={data.get('solped')} | model={data.get('_model')}")
        return jsonify({"success": True, "data": data})

    except Exception as e:
        print(f"[ERROR] {traceback.format_exc()}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def export_excel():
    try:
        body = request.get_json()
        registros = body.get("registros", [])
        if not registros:
            return jsonify({"error": "Sin registros para exportar"}), 400

        buf   = create_excel(registros)
        fname = f"Control_Facturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    print("=" * 60)
    print("  CONTROL DE FACTURAS — Gemini AI (Multi-modelo)")
    print("=" * 60)
    key = CONFIG["api_key"]
    print(f"  API Key  : {key[:8]}...{key[-4:] if key else '???'}")
    print(f"  Modelos  : {' → '.join(CONFIG['models'])}")
    print(f"  Cuentas  : {len(TABLA_CUENTAS)} rubros")
    print(f"  Puerto   : 3000")
    print("=" * 60)
    # threaded=True: Flask atiende múltiples requests simultáneos sin bloquearse
    app.run(host="0.0.0.0", port=3000, debug=False, threaded=True)
